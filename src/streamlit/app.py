"""
Retrieval Ground Truth Labeling App
------------------------------------
Streamlit app untuk anotasi manual kolom label pada kandidat
retrieval ground truth.

Input : data/ground_truth/retrieval_relevant_chunks_candidate_final_fixed.xlsx
Output: data/ground_truth/retrieval_labels_final.xlsx
        data/ground_truth/retrieval_labels_final.csv

Jalankan:
    streamlit run src/streamlit/app.py
    python -m streamlit run src/streamlit/app.py
"""

import html as _html
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components  # kept for declare_component; html replaced by st.iframe
from openpyxl.styles import Alignment, Font, PatternFill

# ── Page setup (MUST be first Streamlit call) ─────────────────────────────────
st.set_page_config(
    page_title="Retrieval Annotator",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent.parent

# Catatan: aplikasi hanya mencoba path v3 di bawah. File tersebut dapat tidak
# tersedia pada checkout baru karena hasil kandidat historis berada di archive;
# bangun atau pulihkan file ini sebelum menjalankan aplikasi anotasi.
# Kandidat aktif: v3 evidence-aware (fresh scan, kolom has_* lengkap)
_CANDIDATE_PRIORITY = [
    ROOT / "data/ground_truth/retrieval_relevant_chunks_candidate_v3_evidence_aware.xlsx",
]
CANDIDATES_XLSX = _CANDIDATE_PRIORITY[0]

_qa_newest      = ROOT / "data/ground_truth/qa_gold_standard_rag_bps_30qa_question_newest.xlsx"
_qa_revised     = ROOT / "data/ground_truth/qa_gold_standard_rag_bps_30qa_fixed_narrative_revised.xlsx"
_qa_fixed       = ROOT / "data/ground_truth/qa_gold_standard_rag_bps_30qa_fixed.xlsx"
_qa_orig        = ROOT / "data/ground_truth/qa_gold_standard_rag_bps_30qa.xlsx"
QA_GOLD_XLSX    = _qa_newest  if _qa_newest.exists()  else (
                  _qa_revised if _qa_revised.exists() else (
                  _qa_fixed   if _qa_fixed.exists()   else _qa_orig))
OUTPUT_XLSX     = ROOT / "data/ground_truth/retrieval_labels_final.xlsx"
OUTPUT_CSV      = ROOT / "data/ground_truth/retrieval_labels_final.csv"

# V2 optional columns
_V2_COLS = ["suggested_label", "confidence", "match_type", "reason", "evidence_quote"]

METHOD_ORDER = {"element_based": 0, "maxmin_semantic": 1, "recursive": 2}

# ── CSS ───────────────────────────────────────────────────────────────────────
CSS = """
<style>
/* ── Base ── */
.stApp { background: #f8fafc; }
section[data-testid="stSidebar"] { background: #ffffff; border-right: 1px solid #e2e8f0; }
footer, header { display: none !important; }

/* ── Query card ── */
.qcard {
    background: white;
    border-radius: 12px;
    padding: 20px 22px;
    margin-bottom: 14px;
    border-left: 4px solid #6366f1;
    box-shadow: 0 1px 4px rgba(0,0,0,0.07);
}
.qcard-title {
    font-size: 17px; font-weight: 700; color: #0f172a; margin-bottom: 4px;
}
.qcard-question {
    font-size: 15px; font-weight: 500; color: #1e293b;
    line-height: 1.55; margin-bottom: 12px;
}
.qcard-gold {
    background: #f0fdf4; border-left: 3px solid #10b981;
    padding: 8px 12px; border-radius: 6px;
    font-size: 14px; color: #166534; line-height: 1.5;
}

/* ── Meta grid ── */
.meta-label { font-size: 10px; font-weight: 700; color: #94a3b8;
              text-transform: uppercase; letter-spacing: .06em; margin-bottom: 1px; }
.meta-value { font-size: 13px; color: #1e293b; margin-bottom: 8px; }

/* ── Badge ── */
.badge {
    display: inline-block; padding: 2px 9px; border-radius: 99px;
    font-size: 11px; font-weight: 600;
}
.b-et-table  { background: #dbeafe; color: #1d4ed8; }
.b-et-para   { background: #ede9fe; color: #6d28d9; }
.b-et-other  { background: #e2e8f0; color: #475569; }

/* ── Chunk card ── */
.ccard {
    background: white; border-radius: 10px;
    padding: 14px 18px; margin-bottom: 10px;
    border: 1.5px solid #e2e8f0;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.ccard-active { border-color: #6366f1; box-shadow: 0 0 0 3px rgba(99,102,241,.12); }
.ccard-lbl-2  { border-left: 4px solid #10b981; }
.ccard-lbl-1  { border-left: 4px solid #f59e0b; }
.ccard-lbl-0  { border-left: 4px solid #ef4444; }
.ccard-lbl-nr { border-left: 4px solid #8b5cf6; }

/* ── Excerpt ── */
.excerpt {
    background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px;
    padding: 10px 13px; margin: 8px 0;
    font-family: ui-monospace, SFMono-Regular, monospace;
    font-size: 12.5px; line-height: 1.65;
    white-space: pre-wrap; word-break: break-word;
    max-height: 110px; overflow-y: auto;
}

/* ── Evidence flags ── */
.flag-row { display:flex; gap:8px; flex-wrap:wrap; margin:6px 0 4px; }
.flag-found  { display:inline-flex; align-items:center; gap:3px; background:#dcfce7;
               color:#166534; border-radius:4px; padding:1px 7px; font-size:11px; font-weight:600; }
.flag-missing{ display:inline-flex; align-items:center; gap:3px; background:#fee2e2;
               color:#991b1b; border-radius:4px; padding:1px 7px; font-size:11px; font-weight:600; }
.flag-na     { display:inline-flex; align-items:center; gap:3px; background:#f1f5f9;
               color:#64748b; border-radius:4px; padding:1px 7px; font-size:11px; }

/* ── Group summary ── */
.grp-summary { background:#f0f9ff; border:1px solid #bae6fd; border-radius:8px;
               padding:8px 14px; margin-bottom:10px; font-size:12px; color:#0c4a6e; }

/* ── Warning box ── */
.warn-box { background:#fffbeb; border-left:4px solid #f59e0b; border-radius:4px;
            padding:7px 12px; font-size:12px; color:#78350f; margin:6px 0; }

/* ── Label badges ── */
.lb { display: inline-block; padding: 2px 8px; border-radius: 99px; font-size: 11px; font-weight: 600; }
.lb-2  { background: #d1fae5; color: #065f46; }
.lb-1  { background: #fef3c7; color: #92400e; }
.lb-0  { background: #fee2e2; color: #991b1b; }
.lb-nr { background: #ede9fe; color: #5b21b6; }
.lb-empty { background: #f1f5f9; color: #94a3b8; }

/* ── Rationale ── */
.rationale {
    background: #eff6ff; border-radius: 4px;
    padding: 5px 10px; font-size: 12px; color: #1d4ed8; margin: 4px 0;
}

/* ── Nav bar ── */
.navbar {
    background: white; border-radius: 10px;
    padding: 10px 16px; margin-bottom: 12px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    display: flex; align-items: center; justify-content: center; gap: 14px;
}

/* ── Kbd ── */
kbd {
    display: inline-block; padding: 1px 6px;
    background: #f1f5f9; border: 1px solid #cbd5e1;
    border-radius: 4px; font-family: monospace; font-size: 11px; color: #334155;
}

/* ── Rule box ── */
.rulebox {
    background: #fafafa; border: 1px solid #e2e8f0;
    border-radius: 8px; padding: 12px 14px; font-size: 12px; line-height: 1.6;
}
</style>
"""


# ── Data loading ──────────────────────────────────────────────────────────────

@st.cache_data(show_spinner="Memuat data kandidat...")
def _load_fresh(_mtime_c: float, _mtime_q: float) -> pd.DataFrame:
    """Load and merge candidates + QA gold. Supports v1 and v2 candidate files."""
    df_c = pd.read_excel(str(CANDIDATES_XLSX), sheet_name="candidates", dtype=str)
    df_q = pd.read_excel(str(QA_GOLD_XLSX), sheet_name="qa_gold", dtype=str)
    df_c = df_c.fillna("")
    df_q = df_q.fillna("")

    qa_cols = [
        "query_id", "source_file", "question", "gold_answer",
        "evidence_text", "table_id", "row_label", "column_label",
        "unit", "gold_value", "evidence_anchor", "evidence_page_pdf",
    ]
    df_q = df_q[[c for c in qa_cols if c in df_q.columns]]

    df = df_c.merge(df_q, on="query_id", how="left")
    df["label"] = df["label"].replace({"None": "", "nan": ""})
    df["annotator"]   = df["annotator"].replace({"None": "", "nan": ""})

    # Auto-excerpt: if chunk_text_excerpt is missing/empty but chunk_text exists
    if "chunk_text" in df.columns:
        if "chunk_text_excerpt" not in df.columns:
            df["chunk_text_excerpt"] = ""
        mask = df["chunk_text_excerpt"].str.strip() == ""
        df.loc[mask, "chunk_text_excerpt"] = (
            df.loc[mask, "chunk_text"].str[:800].str.replace("\n", " ", regex=False)
        )

    # Ensure v2 optional columns exist (empty string if absent)
    for col in _V2_COLS:
        if col not in df.columns:
            df[col] = ""

    if "rationale" not in df.columns:
        df["rationale"] = ""

    return df


def load_data() -> pd.DataFrame:
    """Load fresh or resume from saved output."""
    _mtime_c = CANDIDATES_XLSX.stat().st_mtime if CANDIDATES_XLSX.exists() else 0.0
    _mtime_q = QA_GOLD_XLSX.stat().st_mtime if QA_GOLD_XLSX.exists() else 0.0
    df_fresh = _load_fresh(_mtime_c, _mtime_q)
    if OUTPUT_XLSX.exists():
        try:
            df_saved = pd.read_excel(str(OUTPUT_XLSX), sheet_name="labels", dtype=str)
            df_saved = df_saved.fillna("")
            keys = ["query_id", "method", "chunk_id"]
            if all(k in df_saved.columns for k in keys):
                _restore = [c for c in ["label", "annotator", "rationale"] if c in df_saved.columns]
                saved_map = df_saved.set_index(keys)[_restore]
                idx = df_fresh.set_index(keys)
                for col in _restore:
                    if col not in idx.columns:
                        idx[col] = ""
                    idx[col] = saved_map[col].reindex(idx.index).fillna(idx[col])
                return idx.reset_index()
        except Exception:
            pass
    return df_fresh


# ── Save / export ─────────────────────────────────────────────────────────────

def save_data(df: pd.DataFrame) -> None:
    """Auto-save labels to output XLSX and CSV (silent, called on every label change)."""
    try:
        if len(df) < 10:
            return
        if OUTPUT_XLSX.exists():
            try:
                _existing = pd.read_excel(str(OUTPUT_XLSX), sheet_name="labels", dtype=str)
                if len(_existing) > len(df):
                    return
            except Exception:
                pass
        out_cols = [
            "query_id", "doc_id", "question_preview", "evidence_type", "method",
            "chunk_id", "label", "strength_score", "rationale",
            "chunk_page_start", "chunk_page_end", "chunk_text_excerpt",
            "annotator", "status",
        ]
        df_out = df[[c for c in out_cols if c in df.columns]].copy()
        df_out["status"] = df_out["label"].apply(
            lambda x: "labeled" if x.strip() not in ("", "None") else "needs_manual_validation"
        )
        df_out.to_csv(str(OUTPUT_CSV), index=False, encoding="utf-8")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "labels"

        h_font = Font(bold=True, color="FFFFFF")
        h_fill = PatternFill("solid", fgColor="6366F1")
        fills  = {
            "1": PatternFill("solid", fgColor="D1FAE5"),
            "0": PatternFill("solid", fgColor="FEE2E2"),
            "needs_review": PatternFill("solid", fgColor="EDE9FE"),
        }
        empty_fill = PatternFill("solid", fgColor="F8FAFC")

        ws.append(list(df_out.columns))
        for cell in ws[1]:
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = Alignment(horizontal="center")

        for _, row in df_out.iterrows():
            ws.append(list(row))
            lbl = str(row.get("label", "")).strip()
            fill = fills.get(lbl, empty_fill)
            for cell in ws[ws.max_row]:
                cell.fill = fill

        col_w = {"rationale": 55, "chunk_text_excerpt": 65, "question_preview": 45,
                  "chunk_id": 12, "method": 18, "status": 22, "label": 14}
        for i, name in enumerate(df_out.columns, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = col_w.get(name, 14)

        wb.save(str(OUTPUT_XLSX))
        st.session_state.last_saved = datetime.now()
    except Exception as e:
        st.warning(f"Auto-save gagal: {e}")


# ── Highlight ─────────────────────────────────────────────────────────────────

_STOP_ID = {
    "yang", "dan", "atau", "dari", "pada", "untuk", "dengan", "adalah",
    "dalam", "ini", "itu", "tidak", "oleh", "juga", "akan", "telah",
    "bahwa", "serta", "terhadap", "sebagai", "dapat", "lebih", "yaitu",
    "tersebut", "berdasarkan", "sesuai", "ke", "di", "the", "of", "in",
    "to", "and", "or", "for", "a", "is", "are", "by", "as", "an",
}


def _kws(text: str, n: int = 5) -> list:
    seen, out = set(), []
    for w in re.findall(r"\b\w{4,}\b", text.lower()):
        if w not in _STOP_ID and w not in seen:
            seen.add(w); out.append(w)
        if len(out) >= n:
            break
    return out


def highlight_excerpt(text: str, row: pd.Series) -> str:
    if not text or text.strip() == "":
        return "<em style='color:#94a3b8'>kosong</em>"

    gold_val = str(row.get("gold_value",   "") or "").strip()
    row_lbl  = str(row.get("row_label",    "") or "").strip()
    col_lbl  = str(row.get("column_label", "") or "").strip()
    ev_text  = str(row.get("evidence_text","") or "").strip()

    # (term, background_hex, text_hex, border_hex)
    terms = []
    if gold_val  and gold_val  not in ("None", ""):
        terms.append((gold_val,  "#10b981", "#065f46", "#10b981"))
    if row_lbl   and row_lbl   not in ("None", ""):
        terms.append((row_lbl,   "#3b82f6", "#1e40af", "#3b82f6"))
    if col_lbl   and col_lbl   not in ("None", ""):
        terms.append((col_lbl,   "#f59e0b", "#78350f", "#f59e0b"))
    for kw in _kws(ev_text):
        terms.append((kw, "#8b5cf6", "#4c1d95", "#8b5cf6"))

    # Build spans list
    spans = []
    for term, bg, fg, border in terms:
        if len(term) < 2:
            continue
        for m in re.finditer(re.escape(term), text, re.IGNORECASE):
            spans.append((m.start(), m.end(), bg, fg, border))

    if not spans:
        return f'<span style="white-space:pre-wrap">{_html.escape(text)}</span>'

    spans.sort()
    merged = []
    for s in spans:
        if merged and s[0] < merged[-1][1]:
            continue
        merged.append(s)

    parts, prev = [], 0
    for start, end, bg, fg, border in merged:
        parts.append(_html.escape(text[prev:start]))
        parts.append(
            f'<mark style="background:{bg}22;color:{fg};'
            f'border-bottom:2px solid {border};border-radius:2px;padding:0 2px">'
            f'{_html.escape(text[start:end])}</mark>'
        )
        prev = end
    parts.append(_html.escape(text[prev:]))
    return f'<span style="white-space:pre-wrap">{"".join(parts)}</span>'


# ── Evidence flag helpers ────────────────────────────────────────────────────

def _flag_bool(row: pd.Series, col: str) -> bool:
    """Read has_* flag as bool regardless of string/bool storage."""
    v = str(row.get(col, "") or "").strip()
    return v.lower() in ("true", "1", "yes")


def evidence_flags_html(row: pd.Series) -> str:
    """One-line flag summary for a chunk card."""
    et = str(row.get("evidence_type", "") or "")
    is_table = "table" in et

    def _flag(label: str, found: bool | None) -> str:
        if found is None:
            return f'<span class="flag-na">{label}: N/A</span>'
        return (
            f'<span class="flag-found">✓ {label}</span>'
            if found
            else f'<span class="flag-missing">✗ {label}</span>'
        )

    # Page match: empty page info → N/A
    pg_s = str(row.get("chunk_page_start", "") or "")
    page_val = _flag_bool(row, "page_match")
    page_flag = _flag("page", page_val if pg_s else None)

    flags = []
    if is_table:
        flags.append(_flag("gold_value", _flag_bool(row, "has_gold_value")))
        flags.append(_flag("row_label",  _flag_bool(row, "has_row_label")))
        flags.append(_flag("col_label",  _flag_bool(row, "has_column_label")))
        flags.append(_flag("table_id",   _flag_bool(row, "has_table_id")))
        flags.append(_flag("anchor",     _flag_bool(row, "has_evidence_anchor")))
    else:
        flags.append(_flag("ev_text",    _flag_bool(row, "has_evidence_text")))
        flags.append(_flag("anchor",     _flag_bool(row, "has_evidence_anchor")))
    flags.append(page_flag)
    return '<div class="flag-row">' + "".join(flags) + "</div>"


def group_summary_html(group_df: pd.DataFrame) -> str:
    """Small summary bar shown above the chunk list."""
    n    = len(group_df)
    mts  = group_df["match_type"].tolist() if "match_type" in group_df.columns else []
    n_ex = sum(1 for m in mts if m in ("exact_table_evidence", "exact_narrative_evidence"))
    n_pt = sum(1 for m in mts if "partial" in m)
    n_kw = sum(1 for m in mts if m == "keyword_only")
    has_ex = n_ex > 0
    ex_txt = (
        "<b style='color:#166534'>✓ Ada exact evidence</b>"
        if has_ex
        else "<span style='color:#92400e'>✗ Tidak ada exact evidence</span>"
    )
    return (
        f'<div class="grp-summary">'
        f"{n} kandidat &nbsp;·&nbsp; "
        f"<b>{n_ex}</b> exact &nbsp;·&nbsp; "
        f"<b>{n_pt}</b> partial &nbsp;·&nbsp; "
        f"<b>{n_kw}</b> keyword-only &nbsp;&nbsp;|&nbsp;&nbsp; {ex_txt}"
        f"</div>"
    )


# ── Evidence match type ───────────────────────────────────────────────────────

def evidence_match_type(rationale: str) -> str:
    r = rationale.lower()
    types = []
    if "evidence_text=found" in r:           types.append("evidence_text")
    if "gold_value=" in r:                   types.append("gold_value")
    if "gold_answer_kw=found" in r:          types.append("gold_answer_kw")
    if "anchor_kw=found" in r:               types.append("anchor_kw")
    if "question_kw=found" in r:             types.append("question_kw")
    if "table_id=" in r:                     types.append("table_id")
    if "row_label=" in r:                    types.append("row_label")
    if "col_label=" in r:                    types.append("col_label")
    if "page_" in r and "=match" in r:       types.append("page_match")
    return ", ".join(types) if types else "no_signal"


# ── Label helpers ─────────────────────────────────────────────────────────────

def label_badge_html(label: str) -> str:
    MAP = {
        "1":            '<span class="lb lb-1">● 1 Relevan</span>',
        "0":            '<span class="lb lb-0">● 0 Tidak Relevan</span>',
        "needs_review": '<span class="lb lb-nr">● Needs Review</span>',
    }
    return MAP.get(label, '<span class="lb lb-empty">○ Belum dilabeli</span>')


def apply_label(qid: str, method: str, chunk_id: str, label: str) -> None:
    df = st.session_state.df
    mask = (
        (df["query_id"] == qid)
        & (df["method"]   == method)
        & (df["chunk_id"] == chunk_id)
    )
    df.loc[mask, "label"] = label
    df.loc[mask, "annotator"]   = st.session_state.get("annotator_name", "")
    st.session_state.df = df
    save_data(df)


# ── Keyboard listener (JS) ────────────────────────────────────────────────────

def inject_keyboard_listener() -> None:
    st.iframe(
        """
        <script>
        (function() {
            if (window.parent._kbAnnotAdded) return;
            window.parent._kbAnnotAdded = true;
            window.parent.document.addEventListener('keydown', function(e) {
                var el = window.parent.document.activeElement;
                if (el && ['INPUT','TEXTAREA','SELECT'].includes(el.tagName)) return;
                var map = {'0':'0','1':'1','n':'n','N':'n'};
                if (map[e.key] === undefined) return;
                e.preventDefault();
                var url = new URL(window.parent.location.href);
                url.searchParams.set('_kb', map[e.key]);
                url.searchParams.set('_ts', Date.now().toString());
                window.parent.location.href = url.toString();
            });
        })();
        </script>
        """,
        height=1,
    )


# ── Progress ──────────────────────────────────────────────────────────────────

def get_progress(df: pd.DataFrame) -> dict:
    total   = len(df)
    labeled = int((df["label"].str.strip() != "").sum())
    return {
        "total":     total,
        "labeled":   labeled,
        "unlabeled": total - labeled,
        "pct":       int(labeled / max(total, 1) * 100),
        "n1":        int((df["label"] == "1").sum()),
        "n0":        int((df["label"] == "0").sum()),
        "nnr":       int((df["label"] == "needs_review").sum()),
    }


# ── Session state ─────────────────────────────────────────────────────────────

def init_state(df: pd.DataFrame) -> None:
    if "df"             not in st.session_state: st.session_state.df = df
    if "group_idx"      not in st.session_state:
        # Restore from URL param 'g' so refresh keeps position
        try:
            st.session_state.group_idx = max(0, int(st.query_params.get("g", 0)))
        except (ValueError, TypeError):
            st.session_state.group_idx = 0
    if "last_kb_ts"     not in st.session_state: st.session_state.last_kb_ts = ""
    if "annotator_name" not in st.session_state:
        names = df["annotator"].dropna().astype(str)
        names = names[(names.str.strip() != "") & (names != "annotator")]
        st.session_state.annotator_name = names.mode().iloc[0] if len(names) > 0 else ""
    if "last_saved"     not in st.session_state: st.session_state.last_saved = None
    if "scroll_top"     not in st.session_state: st.session_state.scroll_top = False


# ── Groups ────────────────────────────────────────────────────────────────────

def get_groups(df: pd.DataFrame, filters: dict) -> list:
    fdf = df.copy()
    if filters.get("qid") and filters["qid"] != "All":
        fdf = fdf[fdf["query_id"] == filters["qid"]]
    if filters.get("method") and filters["method"] != "All":
        fdf = fdf[fdf["method"] == filters["method"]]
    if filters.get("only_unlabeled"):
        fdf = fdf[fdf["label"].str.strip() == ""]

    grps = list(fdf.groupby(["query_id", "method"], sort=False).groups.keys())
    grps.sort(key=lambda g: (g[0], METHOD_ORDER.get(g[1], 9)))
    return grps


# ── Sidebar ───────────────────────────────────────────────────────────────────

def render_sidebar(df: pd.DataFrame) -> dict:
    with st.sidebar:
        st.markdown("## 🏷️ Retrieval Annotator")
        st.caption("Validasi manual kandidat retrieval ground truth")
        st.divider()

        # Active candidate file info
        _fname = CANDIDATES_XLSX.name
        _is_v3 = "v3" in _fname.lower()
        _is_v2 = "v2" in _fname.lower() or "merged" in _fname.lower()
        _badge_text  = "v3 evidence-aware" if _is_v3 else ("v2" if _is_v2 else "v1")
        _badge_color = "#d1fae5" if _is_v3 else ("#fef3c7" if _is_v2 else "#fee2e2")
        _n_total  = len(df)
        _n_qids   = df["query_id"].nunique()
        _n_meths  = df["method"].nunique() if "method" in df.columns else 0
        _n_groups = df.groupby(["query_id", "method"]).ngroups if "method" in df.columns else 0
        st.markdown(
            f"<div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;"
            f"padding:8px 10px;font-size:11px;line-height:1.8;margin-bottom:4px'>"
            f"<b>Kandidat aktif</b>&nbsp;"
            f"<span style='background:{_badge_color};padding:1px 7px;border-radius:99px;"
            f"font-size:10px;font-weight:700'>{_badge_text}</span><br>"
            f"<b>{_n_total}</b> baris &nbsp;·&nbsp; <b>{_n_qids}</b> query &nbsp;·&nbsp; "
            f"<b>{_n_meths}</b> metode &nbsp;·&nbsp; <b>{_n_groups}</b> grup"
            f"</div>",
            unsafe_allow_html=True,
        )
        st.caption(str(CANDIDATES_XLSX))

        # Progress
        p = get_progress(df)
        st.markdown(
            f"<div style='text-align:center;padding:6px 0'>"
            f"<div style='font-size:36px;font-weight:800;color:#6366f1'>{p['pct']}%</div>"
            f"<div style='color:#94a3b8;font-size:11px;margin-top:-4px'>progress anotasi</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
        st.progress(p["pct"] / 100)

        c1, c2 = st.columns(2)
        c1.metric("Dilabeli",  p["labeled"])
        c2.metric("Belum",     p["unlabeled"])

        st.markdown(
            f'<div style="display:flex;gap:5px;flex-wrap:wrap;margin:6px 0 2px">'
            f'<span class="lb lb-1">1: {p["n1"]}</span>'
            f'<span class="lb lb-0">0: {p["n0"]}</span>'
            f'<span class="lb lb-nr">NR: {p["nnr"]}</span>'
            f"</div>",
            unsafe_allow_html=True,
        )

        st.divider()

        # Annotator
        name = st.text_input(
            "Nama Anotator",
            value=st.session_state.annotator_name,
            placeholder="inisial / nama...",
        )
        st.session_state.annotator_name = name

        st.divider()

        # Filters
        st.markdown("**Filter**")
        all_qids = ["All"] + sorted(df["query_id"].unique().tolist())
        qid_f    = st.selectbox("Query ID", all_qids, key="f_qid")
        meth_f   = st.selectbox("Method",
                                 ["All", "element_based", "maxmin_semantic", "recursive"],
                                 key="f_method")
        only_unl = st.checkbox("Hanya belum dilabeli", key="f_unl")

        st.divider()

        # Rules
        st.markdown("**Aturan Anotasi**")
        st.markdown(
            """<div class="rulebox">
            <b>Label 1 — Relevan</b><br>
            <small>• <b>Tabel</b>: chunk memuat row_label + col_label + gold_value secara jelas, ATAU relevan sebagian.<br>
            • <b>Narasi</b>: chunk memuat klaim yang relevan dengan question, baik sebagai bukti utama maupun konteks pendukung.<br>
            • ⚠ Jangan beri 1 hanya karena ada keyword atau angka tanpa konteks.</small>
            <br><br>
            <b>Label 0 — Tidak Relevan</b><br>
            <small>Chunk tidak memuat bukti atau konteks relevan untuk menjawab question.</small>
            </div>""",
            unsafe_allow_html=True,
        )

        st.divider()

        # Keyboard shortcuts
        st.markdown("**Keyboard Shortcuts**")
        st.markdown(
            "<div style='font-size:12px;line-height:2.2'>"
            "<kbd>1</kbd> Relevan &nbsp; <kbd>0</kbd> Tidak Relevan<br>"
            "<kbd>N</kbd> Next grup<br>"
            "<span style='color:#94a3b8;font-size:11px'>⚠ Shortcut berlaku untuk kandidat"
            " <b>pertama yang belum dilabeli</b> dalam grup aktif.</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        st.divider()

        # ── Save / Export ────────────────────────────────────────────────
        _last = st.session_state.get("last_saved", None)
        _last_str = _last.strftime("%H:%M:%S") if _last else "belum disimpan"
        st.markdown(
            f"<div style='font-size:11px;color:#64748b;margin-bottom:4px'>"
            f"Terakhir disimpan: <b>{_last_str}</b></div>",
            unsafe_allow_html=True,
        )
        if st.button("💾 Simpan Progress", type="primary", use_container_width=True):
            save_data(st.session_state.df)
            st.success("✓ Progress tersimpan!")
        if st.button("📥 Export Final", use_container_width=True):
            save_data(st.session_state.df)
            st.success(f"Tersimpan ke {OUTPUT_XLSX.name}")

    return {"qid": qid_f, "method": meth_f, "only_unlabeled": only_unl}


# ── Query info panel ──────────────────────────────────────────────────────────

def render_query_panel(row: pd.Series) -> None:
    et = str(row.get("evidence_type", ""))
    et_cls = "b-et-table" if "table" in et else ("b-et-para" if "para" in et else "b-et-other")

    qid      = str(row.get("query_id", ""))
    doc_id   = str(row.get("doc_id", ""))
    source   = str(row.get("source_file", ""))
    question = str(row.get("question", row.get("question_preview", "")))
    gold_ans = str(row.get("gold_answer", ""))
    ev_text  = str(row.get("evidence_text", ""))
    anchor   = str(row.get("evidence_anchor", ""))
    table_id = str(row.get("table_id", ""))
    row_lbl  = str(row.get("row_label", ""))
    col_lbl  = str(row.get("column_label", ""))
    unit     = str(row.get("unit", ""))
    gold_val = str(row.get("gold_value", ""))

    st.markdown(
        f"""<div class="qcard">
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
            <span style="font-size:17px;font-weight:700;color:#0f172a">{qid}</span>
            <span style="background:#f1f5f9;padding:2px 8px;border-radius:6px;
                         font-size:11px;color:#475569">{doc_id}</span>
            <span class="badge {et_cls}">{et}</span>
        </div>
        <div style="font-size:11px;color:#94a3b8;margin-bottom:2px">PERTANYAAN</div>
        <div class="qcard-question">{_html.escape(question)}</div>
        <div style="font-size:11px;color:#94a3b8;margin-bottom:4px">JAWABAN GOLD</div>
        <div class="qcard-gold">{_html.escape(gold_ans)}</div>
        </div>""",
        unsafe_allow_html=True,
    )

    # Meta grid
    meta = [
        ("Evidence Text", ev_text),
        ("Evidence Anchor", anchor),
        ("Source File",    source),
    ]
    if "table" in et:
        meta += [
            ("Table ID",      table_id),
            ("Row Label",     row_lbl),
            ("Column Label",  col_lbl),
            ("Unit",          unit),
            ("Gold Value",    gold_val),
        ]

    valid = [(k, v) for k, v in meta if v and v not in ("", "None")]
    if valid:
        cols = st.columns(min(len(valid), 4))
        for i, (lbl, val) in enumerate(valid):
            with cols[i % len(cols)]:
                st.markdown(
                    f'<div class="meta-label">{lbl}</div>'
                    f'<div class="meta-value">{_html.escape(str(val))}</div>',
                    unsafe_allow_html=True,
                )


# ── Chunk cards ───────────────────────────────────────────────────────────────

_MT_SORT = {
    "exact_table_evidence":    0, "exact_narrative_evidence": 1,
    "partial_table_evidence":  2, "partial_narrative_context": 3,
    "keyword_only":            4, "not_found":                 5,
}
_CONF_SORT = {"high": 0, "medium": 1, "low": 2}


def sort_group_df(gdf: pd.DataFrame) -> pd.DataFrame:
    """Sort candidates within a group by evidence quality."""
    if gdf.empty:
        return gdf
    tmp = gdf.copy()
    tmp["_mt"]   = tmp.get("match_type",      pd.Series(dtype=str)).map(_MT_SORT).fillna(6)
    tmp["_sl"]   = tmp.get("suggested_label", pd.Series(dtype=str)).apply(
        lambda x: -int(x) if str(x).strip().lstrip("-").isdigit() else 0)
    tmp["_conf"] = tmp.get("confidence",      pd.Series(dtype=str)).map(_CONF_SORT).fillna(3)
    tmp["_sc"]   = tmp.get("strength_score",  pd.Series(dtype=str)).apply(
        lambda x: -float(x) if str(x).replace(".", "").lstrip("-").isdigit() else 0)
    return (tmp.sort_values(["_mt", "_sl", "_conf", "_sc"])
               .drop(columns=["_mt", "_sl", "_conf", "_sc"])
               .reset_index(drop=True))


def render_chunk_card(row: pd.Series, active: bool) -> None:
    """Render a single chunk card with label buttons."""
    label     = str(row.get("label", "") or "").strip()
    qid       = str(row.get("query_id", ""))
    method    = str(row.get("method", ""))
    chunk_id  = str(row.get("chunk_id", ""))
    rationale = str(row.get("rationale", ""))
    excerpt   = str(row.get("chunk_text_excerpt", ""))
    full_text = str(row.get("chunk_text", ""))
    strength  = str(row.get("strength_score", ""))
    pg_start  = str(row.get("chunk_page_start", ""))
    pg_end    = str(row.get("chunk_page_end", ""))
    et        = str(row.get("evidence_type", "") or "")
    pages     = pg_start if (not pg_end or pg_end == pg_start) else f"{pg_start}\u2013{pg_end}"
    match_type = evidence_match_type(rationale)

    # V2 optional fields
    suggested  = str(row.get("suggested_label", "") or "").strip()
    confidence = str(row.get("confidence",      "") or "").strip()
    mt_v2      = str(row.get("match_type",      "") or "").strip()
    ev_quote   = str(row.get("evidence_quote",  "") or "").strip()
    reason_v2  = str(row.get("reason",          "") or "").strip()

    lbl_cls = {"2": "ccard-lbl-2", "1": "ccard-lbl-1", "0": "ccard-lbl-0",
               "needs_review": "ccard-lbl-nr"}.get(label, "")
    act_cls = "ccard-active" if active else ""
    pg_badge = (
        f"<span style='background:#fef3c7;padding:1px 7px;border-radius:4px;"
        f"font-size:11px;color:#92400e'>pg {pages}</span>"
        if pages else ""
    )

    hi = highlight_excerpt(excerpt, row)

    _slbl_colors = {"2": ("#d1fae5","#065f46"), "1": ("#fef3c7","#92400e"), "0": ("#fee2e2","#991b1b")}
    _sc_c, _stc = _slbl_colors.get(suggested, ("#f1f5f9", "#64748b"))
    _suggested_badge = (
        f"<span style='background:{_sc_c};color:{_stc};padding:1px 7px;border-radius:4px;"
        f"font-size:11px;font-weight:600'>saran:{suggested}</span>"
        if suggested else ""
    )
    _conf_colors = {"high": "#d1fae5", "medium": "#fef3c7", "low": "#fee2e2"}
    _conf_badge = (
        f"<span style='background:{_conf_colors.get(confidence, '#f1f5f9')};padding:1px 6px;"
        f"border-radius:4px;font-size:10px;color:#475569'>{confidence}</span>"
        if confidence else ""
    )
    display_mt = mt_v2 if mt_v2 else match_type
    _pg = f" {pg_badge}" if pg_badge else ""
    _flags = evidence_flags_html(row)

    card_html = (
        f'<div class="ccard {act_cls} {lbl_cls}">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">'
        f'<div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap">'
        f'<span style="font-size:12px;font-weight:600;color:#475569">CHUNK #{chunk_id}</span>'
        f'<span style="background:#f1f5f9;padding:1px 7px;border-radius:4px;font-size:11px;color:#64748b">{method}</span>'
        f'{_pg}'
        f'<span style="background:#e0f2fe;padding:1px 7px;border-radius:4px;font-size:11px;color:#0369a1">str={strength}</span>'
        f'{_suggested_badge}{_conf_badge}'
        f'</div>'
        f'{label_badge_html(label)}'
        f'</div>'
        f'<div style="font-size:11px;color:#64748b;margin:2px 0 4px">'
        f'match: <b style="color:#334155">{_html.escape(display_mt)}</b>'
        f'</div>'
        f'{_flags}'
        f'<div class="excerpt">{hi}</div>'
        f'</div>'
    )
    st.markdown(card_html, unsafe_allow_html=True)

    # Full chunk expander
    if full_text:
        with st.expander("Show full chunk", expanded=False):
            st.text(full_text)

    # Evidence details expander
    with st.expander("Why this candidate?", expanded=False):
        st.markdown(
            f"**match_type:** `{_html.escape(display_mt)}`  \n"
            f"**suggested_label:** `{suggested or '—'}`  \n"
            f"**confidence:** `{confidence or '—'}`"
        )
        st.markdown("**reason:**")
        st.code(reason_v2 if reason_v2 else "—", language=None)
        st.markdown("**evidence_quote:**")
        if ev_quote:
            st.markdown(
                f'<div style="background:#eff6ff;border-left:3px solid #3b82f6;border-radius:4px;'
                f'padding:8px 12px;font-size:12px;color:#1e40af;white-space:pre-wrap">'
                f'{_html.escape(ev_quote)}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.caption("No evidence quote available")

    # Label buttons + rationale editor
    b1, b2, b3, spacer, rat_col = st.columns([1.3, 1.3, 1.1, 0.2, 3])
    key_sfx = f"{qid}_{method}_{chunk_id}"

    with b1:
        if st.button(
            "1 \u2014 Relevan", key=f"b1_{key_sfx}",
            type="primary" if label == "1" else "secondary",
            use_container_width=True,
        ):
            apply_label(qid, method, chunk_id, "1")
            st.rerun()

    with b2:
        if st.button(
            "0 \u2014 Tidak Relevan", key=f"b0_{key_sfx}",
            type="primary" if label == "0" else "secondary",
            use_container_width=True,
        ):
            apply_label(qid, method, chunk_id, "0")
            st.rerun()

    with b3:
        if st.button(
            "Review", key=f"br_{key_sfx}",
            type="primary" if label == "needs_review" else "secondary",
            use_container_width=True,
        ):
            apply_label(qid, method, chunk_id, "needs_review")
            st.rerun()

    with rat_col:
        new_rat = st.text_input(
            "Rationale", value=rationale,
            key=f"rat_{key_sfx}",
            label_visibility="collapsed",
            placeholder="Catatan anotator (opsional)...",
        )
        if new_rat != rationale:
            mask = (
                (st.session_state.df["query_id"] == qid)
                & (st.session_state.df["method"]   == method)
                & (st.session_state.df["chunk_id"] == chunk_id)
            )
            st.session_state.df.loc[mask, "rationale"] = new_rat
            save_data(st.session_state.df)

    st.markdown("")  # breathing room


# ── Main ──────────────────────────────────────────────────────────────────────

def _scroll_to_top() -> None:
    """Inject a one-shot JS snippet that scrolls the Streamlit main container to top."""
    st.iframe(
        """
        <script>
        (function() {
            var selectors = [
                'section[data-testid="stMain"]',
                '.main',
                'section.main'
            ];
            function tryScroll() {
                var done = false;
                selectors.forEach(function(sel) {
                    var el = window.parent.document.querySelector(sel);
                    if (el) { el.scrollTop = 0; done = true; }
                });
                if (!done) {
                    window.parent.document.documentElement.scrollTop = 0;
                    window.parent.document.body.scrollTop = 0;
                }
            }
            tryScroll();
            setTimeout(tryScroll, 80);
        })();
        </script>
        """,
        height=1,
    )


def main() -> None:
    st.markdown(CSS, unsafe_allow_html=True)
    inject_keyboard_listener()

    # Load / resume
    if "df" not in st.session_state:
        df = load_data()
        init_state(df)
    else:
        init_state(st.session_state.df)

    # Scroll to top after navigation
    if st.session_state.get("scroll_top", False):
        st.session_state.scroll_top = False
        _scroll_to_top()

    # Sidebar
    filters = render_sidebar(st.session_state.df)

    # Handle keyboard action via URL params
    kb_action = st.query_params.get("_kb", "")
    kb_ts     = st.query_params.get("_ts", "")

    # Filtered groups
    groups = get_groups(st.session_state.df, filters)

    if not groups:
        st.info("Tidak ada data sesuai filter. Ubah filter di sidebar.")
        return

    # Clamp index and sync to URL so refresh restores position
    gidx = min(st.session_state.group_idx, len(groups) - 1)
    st.session_state.group_idx = gidx
    st.query_params["g"] = str(gidx)
    qid, method = groups[gidx]

    # Display number per query_id (sequential, gaps ignored)
    _all_qids = sorted(st.session_state.df["query_id"].unique())
    _qid_no   = {q: i + 1 for i, q in enumerate(_all_qids)}
    _q_disp   = f"Q{_qid_no.get(qid, '?')} / {len(_all_qids)}"

    # Current group chunks
    mask = (
        (st.session_state.df["query_id"] == qid)
        & (st.session_state.df["method"]   == method)
    )
    group_df = sort_group_df(st.session_state.df[mask].copy())

    # Active chunk = first unlabeled
    unlabeled = group_df[group_df["label"].str.strip() == ""].index.tolist()
    active_ci = unlabeled[0] if unlabeled else 0

    # Process keyboard action
    if kb_action and kb_ts != st.session_state.last_kb_ts:
        st.session_state.last_kb_ts = kb_ts
        st.query_params.pop("_kb", None)
        st.query_params.pop("_ts", None)
        if kb_action == "n":
            st.session_state.group_idx = (gidx + 1) % len(groups)
            st.session_state.scroll_top = True
            st.rerun()
        elif kb_action in ("0", "1") and len(group_df) > active_ci:
            r = group_df.iloc[active_ci]
            apply_label(qid, method, r["chunk_id"], kb_action)
            st.rerun()

    # ── Navigation bar ───────────────────────────────────────────────────────
    nc, prev_c, next_c = st.columns([0.13, 0.74, 0.13])
    with prev_c:
        st.markdown(
            f"<div style='text-align:center;background:white;border-radius:10px;"
            f"padding:10px;box-shadow:0 1px 4px rgba(0,0,0,.06)'>"
            f"<span style='font-size:12px;color:#94a3b8'>Grup</span>&nbsp;"
            f"<span style='font-size:17px;font-weight:800;color:#1e293b'>{gidx+1}</span>"
            f"<span style='font-size:12px;color:#94a3b8'> / {len(groups)}</span>"
            f"&nbsp;&nbsp;·&nbsp;&nbsp;"
            f"<span style='font-size:15px;font-weight:700;color:#6366f1'>{_q_disp}</span>"
            f"&nbsp;"
            f"<span style='font-size:12px;color:#94a3b8'>{method}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with nc:
        if st.button("◀", use_container_width=True, disabled=(gidx == 0), key="nav_prev"):
            st.session_state.group_idx = gidx - 1
            st.session_state.scroll_top = True
            st.rerun()
    with next_c:
        if st.button("▶", use_container_width=True,
                      disabled=(gidx >= len(groups) - 1), key="nav_next"):
            st.session_state.group_idx = gidx + 1
            st.session_state.scroll_top = True
            st.rerun()

    st.markdown("")

    # ── Query info ────────────────────────────────────────────────────────────
    if len(group_df) > 0:
        render_query_panel(group_df.iloc[0])

    # ── Chunk cards ───────────────────────────────────────────────────────────
    if len(group_df) > 0:
        st.markdown(group_summary_html(group_df), unsafe_allow_html=True)
    st.markdown(
        f"<div style='font-size:13px;font-weight:600;color:#475569;"
        f"margin-bottom:8px'>CHUNK KANDIDAT "
        f"<span style='color:#94a3b8;font-weight:400'>({len(group_df)} chunks)</span></div>",
        unsafe_allow_html=True,
    )

    if len(group_df) == 0:
        st.info("Tidak ada chunk untuk grup ini.")
    else:
        for i, (_, row) in enumerate(group_df.iterrows()):
            render_chunk_card(row, active=(i == active_ci))

    # ── Bottom navigation bar ─────────────────────────────────────────────────
    st.divider()
    bnc, bprev_c, bnext_c = st.columns([0.13, 0.74, 0.13])
    with bprev_c:
        st.markdown(
            f"<div style='text-align:center;background:white;border-radius:10px;"
            f"padding:8px;box-shadow:0 1px 4px rgba(0,0,0,.06)'>"
            f"<span style='font-size:12px;color:#94a3b8'>Grup</span>&nbsp;"
            f"<span style='font-size:17px;font-weight:800;color:#1e293b'>{gidx+1}</span>"
            f"<span style='font-size:12px;color:#94a3b8'> / {len(groups)}</span>"
            f"&nbsp;&nbsp;·&nbsp;&nbsp;"
            f"<span style='font-size:15px;font-weight:700;color:#6366f1'>{_q_disp}</span>"
            f"&nbsp;<span style='font-size:12px;color:#94a3b8'>{method}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with bnc:
        if st.button("◀", use_container_width=True, disabled=(gidx == 0), key="bot_prev"):
            st.session_state.group_idx = gidx - 1
            st.session_state.scroll_top = True
            st.rerun()
    with bnext_c:
        if st.button("▶", use_container_width=True,
                      disabled=(gidx >= len(groups) - 1), key="bot_next"):
            st.session_state.group_idx = gidx + 1
            st.session_state.scroll_top = True
            st.rerun()

    # ── Jump to group ─────────────────────────────────────────────────────────
    st.divider()
    with st.expander("🔍 Jump ke Query ID", expanded=False):
        jq = st.selectbox("Query ID", sorted(st.session_state.df["query_id"].unique()),
                           key="j_qid")
        jm = st.selectbox("Method", ["element_based", "maxmin_semantic", "recursive"],
                           key="j_method")
        if st.button("Go", key="btn_jump"):
            target = (jq, jm)
            if target in groups:
                st.session_state.group_idx = groups.index(target)
                st.rerun()
            else:
                st.warning("Grup tidak ditemukan dengan filter aktif. Coba reset filter.")


if __name__ == "__main__":
    main()
