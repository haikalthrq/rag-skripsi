"""
Konversi ground truth CSV ke format JSON nested per chunking method.

Input:
  qa_gold_validated.csv      — QA pairs (query_id, question, reference_answer, ...)
  retrieval_labels_filled.csv — Label relevansi chunk per query × method

Output:
  JSON file dengan format:
  [
    {
      "id": "Q001",
      "doc_id": "DOC01_BIK",
      "question": "...",
      "reference_answer": "...",
      "evidence_page": "...",
      "anchor": "...",
      "relevant_chunk_ids": {
        "element_based": ["doc_chunks_embeddings_4"],
        "recursive": ["doc_chunks_embeddings_2"],
        "maxmin_semantic": ["doc_chunks_embeddings_1"]
      }
    },
    ...
  ]

Penggunaan:
  # Binary (threshold=1, label 1 = relevan, label 0 = tidak relevan)
  python scripts/convert_ground_truth_to_json.py \\
      --qa_csv data/ground_truth/qa_gold_standard_rag_bps_30qa_question_newest.xlsx \\
      --labels_csv data/ground_truth/retrieval_labels_final.csv \\
      --output data/ground_truth/qa_pairs_binary.json \\
      --relevance_threshold 1
"""

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# ─── Mapping: doc_id → prefix ChromaDB (nama file embedding tanpa ekstensi) ───

DOC_TO_FILE_STEM = {
    "DOC01_BIK":    "benchmark-indeks-konstruksi--2016-100---2018---2023_chunks_embeddings",
    "DOC02_BSK":    "benchmark-statistik-konstruksi--2018---2023_chunks_embeddings",
    "DOC03_CERDAS": (
        "cerita-data-statistik-untuk-indonesia---mismatch-pendidikan---"
        "pekerjaan-pemuda-indonesia--implikasi-bagi-bonus-demografi_chunks_embeddings"
    ),
    "DOC04_IUV":    "indeks-unit-value-ekspor-impor---agustus-2025_chunks_embeddings",
    "DOC05_LNPRT":  "neraca-lembaga-non-profit-yang-melayani-rumahtangga--2022-2024_chunks_embeddings",
    "DOC06_NPU":    "neraca-pemerintahan-umum-indonesia-2019-2024_chunks_embeddings",
    "DOC07_NRT":    "neraca-rumah-tangga-indonesia--2022-2024_chunks_embeddings",
    "DOC08_PEND":   "statistik-pendidikan-2025_chunks_embeddings",
    "DOC09_IMPOR":  "statistik-perdagangan-luar-negeri-bulanan-impor--agustus-2025_chunks_embeddings",
    "DOC10_MODA":   (
        "statistik-perdagangan-luar-negeri-menurut-moda-transportasi--"
        "2023-dan-2024_chunks_embeddings"
    ),
}

# ─── Mapping: nama metode di CSV → kode pipeline ──────────────────────────────

CSV_METHOD_TO_CODE = {
    "element":        "element_based",
    "element_based":  "element_based",
    "semantic_maxmin":  "maxmin_semantic",
    "maxmin_semantic":  "maxmin_semantic",
    "recursive":      "recursive",
}

ALL_METHODS = ["element_based", "maxmin_semantic", "recursive"]

# ─── Kolom wajib ──────────────────────────────────────────────────────────────

QA_REQUIRED_COLS = {
    "query_id", "doc_id", "question",
}

LABELS_REQUIRED_COLS = {
    "query_id", "doc_id", "method", "chunk_id", "label",
}


def load_qa(path: str, sheet: str = "qa_gold") -> dict:
    """
    Baca QA pairs dari CSV atau xlsx, return dict query_id → row.
    Untuk xlsx, baca sheet `sheet` (default: 'qa_gold').
    """
    p = Path(path)
    rows = {}

    if p.suffix.lower() in (".xlsx", ".xls"):
        if not _OPENPYXL_AVAILABLE:
            print("[ERROR] openpyxl tidak tersedia. Install: pip install openpyxl", file=sys.stderr)
            sys.exit(1)
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            print(f"[ERROR] Sheet '{sheet}' tidak ada di {path}. Tersedia: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
        ws = wb[sheet]
        row_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(row_iter)]
        for values in row_iter:
            row = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(values)}
            if row.get("query_id"):
                rows[row["query_id"]] = row
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            cols = set(reader.fieldnames or [])
            missing = QA_REQUIRED_COLS - cols
            if missing:
                print(f"[ERROR] qa_csv kekurangan kolom: {missing}", file=sys.stderr)
                sys.exit(1)
            for row in reader:
                if row.get("query_id"):
                    rows[row["query_id"]] = row

    print(f"[INFO] Loaded {len(rows)} QA pairs dari {path}")
    return rows


def load_labels(path: str, threshold: int) -> dict:
    """
    Baca retrieval_labels_filled.csv.
    Return: {query_id: {pipeline_method: [chroma_chunk_id, ...]}}
    Hanya include chunk dengan label >= threshold.
    """
    result: dict = defaultdict(lambda: defaultdict(list))
    skipped_doc = set()
    skipped_method = set()
    total = 0
    included = 0

    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = set(reader.fieldnames or [])
        missing = LABELS_REQUIRED_COLS - cols
        if missing:
            print(f"[ERROR] labels_csv kekurangan kolom: {missing}", file=sys.stderr)
            sys.exit(1)

        for row in reader:
            total += 1
            label_str = row.get("label", "").strip()
            try:
                label = int(label_str)
            except ValueError:
                continue

            if label < threshold:
                continue

            q_id   = row["query_id"].strip()
            doc_id = row["doc_id"].strip()
            method_csv = row["method"].strip()
            chunk_id_int = row["chunk_id"].strip()

            # Validasi doc_id
            if doc_id not in DOC_TO_FILE_STEM:
                skipped_doc.add(doc_id)
                continue

            # Validasi method
            if method_csv not in CSV_METHOD_TO_CODE:
                skipped_method.add(method_csv)
                continue

            pipeline_method = CSV_METHOD_TO_CODE[method_csv]
            file_stem = DOC_TO_FILE_STEM[doc_id]
            chroma_id = f"{file_stem}_{chunk_id_int}"

            result[q_id][pipeline_method].append(chroma_id)
            included += 1

    if skipped_doc:
        print(f"[WARN] doc_id tidak dikenali (dilewati): {skipped_doc}", file=sys.stderr)
    if skipped_method:
        print(f"[WARN] method tidak dikenali (dilewati): {skipped_method}", file=sys.stderr)

    print(
        f"[INFO] Labels: {total} baris dibaca, "
        f"{included} chunk masuk (threshold >= {threshold})"
    )
    return dict(result)


def build_output(qa_rows: dict, labels: dict, threshold: int) -> list:
    """Gabungkan QA + labels menjadi list output JSON."""
    output = []
    no_relevant = []

    for q_id in sorted(qa_rows.keys()):
        row = qa_rows[q_id]
        reference = (
            row.get("revised_gold_answer", "").strip()
            or row.get("gold_answer_original", "").strip()
            or row.get("gold_answer", "").strip()
        )

        query_labels = labels.get(q_id, {})
        relevant_chunk_ids = {
            m: query_labels.get(m, []) for m in ALL_METHODS
        }

        total_relevant = sum(len(v) for v in relevant_chunk_ids.values())
        if total_relevant == 0:
            no_relevant.append(q_id)

        entry = {
            "id":               q_id,
            "doc_id":           row.get("doc_id", "").strip(),
            "question":         row.get("question", "").strip(),
            "reference_answer": reference,
            "evidence_page":    row.get("evidence_pdf_page", row.get("evidence_page", "")).strip(),
            "anchor":           row.get("evidence_text", row.get("evidence_anchor", "")).strip(),
            "relevant_chunk_ids": relevant_chunk_ids,
        }
        output.append(entry)

    if no_relevant:
        print(
            f"[WARN] {len(no_relevant)} query tanpa chunk relevan "
            f"(threshold={threshold}): {no_relevant[:5]}{'...' if len(no_relevant) > 5 else ''}",
            file=sys.stderr,
        )

    return output


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Konversi ground truth CSV ke JSON nested per chunking method",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--qa_csv", type=str,
        default="data/ground_truth/qa_gold_standard_rag_bps_30qa_question_newest.xlsx",
        help="Path ke QA gold file (.csv atau .xlsx). Untuk xlsx, baca sheet 'qa_gold'.",
    )
    # Catatan: loader label membaca CSV dengan csv.DictReader, bukan XLSX.
    # Default retrieval_labels_filled.csv adalah nama historis dan mungkin tidak
    # tersedia; checkout saat ini memakai retrieval_labels_final.csv.
    parser.add_argument(
        "--labels_csv", type=str,
        default="data/ground_truth/retrieval_labels_filled.csv",
        help="Path ke retrieval_labels_filled.csv",
    )
    parser.add_argument(
        "--output", type=str, required=True,
        help="Path output JSON (contoh: data/ground_truth/qa_pairs_strict.json)",
    )
    parser.add_argument(
        "--relevance_threshold", type=int, default=1, choices=[0, 1],
        help="Minimum label untuk dianggap relevan (1=relevan, 0=semua). Default: 1",
    )
    args = parser.parse_args()

    qa_rows = load_qa(args.qa_csv)
    labels  = load_labels(args.labels_csv, args.relevance_threshold)
    output  = build_output(qa_rows, labels, args.relevance_threshold)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"[INFO] Saved {len(output)} entries → {out_path}")
    print(f"[INFO] Mode: {'semua relevan (threshold=0)' if args.relevance_threshold == 0 else 'binary (label>=1=relevan)'}")


if __name__ == "__main__":
    main()
