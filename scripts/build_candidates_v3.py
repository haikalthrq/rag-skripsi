"""
build_candidates.py — Evidence-Aware Retrieval Candidate Generator
=========================================================================
Membangun kandidat chunk evidence-aware untuk anotasi manual retrieval
ground truth.
  1. Word-boundary matching untuk row_label / col_label — cegah 'Riau' match
     'Kepulauan Riau' secara salah.
  2. Boolean flag kolom eksplisit: has_gold_value, has_row_label, ...
  3. Pre-filter top-10 per grup → simpan top-5 terbaik.
  4. evidence_type paragraph_table ditangani dengan hybrid scoring.
  5. evidence_anchor dihitung sebagai sinyal tersendiri.
  6. Summary comparison otomatis disimpan ke .txt.
  7. Confidence lebih granular.

Input:
  data/ground_truth/qa_gold_standard_rag_bps_30qa_question_newest.xlsx  (sheet: qa_gold)
  data/chunked/{method}/{doc}_chunks.json

Output:
  data/ground_truth/retrieval_relevant_chunks_candidate_v3_evidence_aware.xlsx
  data/ground_truth/retrieval_relevant_chunks_candidate_v3_evidence_aware.csv
  data/ground_truth/summary_candidate_v3_evidence_aware.txt
  data/ground_truth/validation_candidate_v3_after_narrative_revision.txt

match_type:
  exact_table_evidence       gold_value + row_label + col_label hadir bersama
  partial_table_evidence     sebagian komponen tabel hadir (tapi belum lengkap)
  exact_narrative_evidence   evidence_text ditemukan exact di chunk
  partial_narrative_context  konteks relevan tapi belum cukup sebagai bukti utama
  keyword_only               hanya keyword tersebar, tanpa relasi bukti
  not_relevant               chunk tidak membantu menjawab pertanyaan
  not_found                  tidak ada kandidat relevan ditemukan sama sekali

confidence:
  high    exact evidence, semua sinyal utama hadir
  medium  partial evidence, sebagian sinyal hadir
  low     keyword_only atau ambigu / rawan false positive
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    import pandas as pd
except ImportError as e:
    raise SystemExit(f"[ERROR] Dependensi tidak tersedia: {e}\nInstall: pip install openpyxl pandas")

# ─── Paths ─────────────────────────────────────────────────────────────────────

ROOT      = Path(__file__).resolve().parent.parent
QA_XLSX   = ROOT / "data/ground_truth/qa_gold_standard_rag_bps_30qa_question_newest.xlsx"
CHUNK_DIR = ROOT / "data/chunked"
OUT_DIR   = ROOT / "data/ground_truth"

V2_XLSX   = ROOT / "data/ground_truth/retrieval_relevant_chunks_candidate_v2_top5.xlsx"

OUT_STEM  = "retrieval_relevant_chunks_candidate_v3_evidence_aware"

# ─── Doc → chunk filename mapping ──────────────────────────────────────────────

DOC_MAP = {
    "DOC01_BIK":    "benchmark-indeks-konstruksi--2016-100---2018---2023",
    "DOC02_BSK":    "benchmark-statistik-konstruksi--2018---2023",
    "DOC03_CERDAS": (
        "cerita-data-statistik-untuk-indonesia---mismatch-pendidikan---"
        "pekerjaan-pemuda-indonesia--implikasi-bagi-bonus-demografi"
    ),
    "DOC04_IUV":    "indeks-unit-value-ekspor-impor---agustus-2025",
    "DOC05_LNPRT":  "neraca-lembaga-non-profit-yang-melayani-rumahtangga--2022-2024",
    "DOC06_NPU":    "neraca-pemerintahan-umum-indonesia-2019-2024",
    "DOC07_NRT":    "neraca-rumah-tangga-indonesia--2022-2024",
    "DOC08_PEND":   "statistik-pendidikan-2025",
    "DOC09_IMPOR":  "statistik-perdagangan-luar-negeri-bulanan-impor--agustus-2025",
    "DOC10_MODA":   "statistik-perdagangan-luar-negeri-menurut-moda-transportasi--2023-dan-2024",
}

METHODS = ["element_based", "maxmin_semantic", "recursive"]

PRE_K_DEFAULT = 10
TOP_K_DEFAULT = 5

# ─── Compound-entity prefix guard ──────────────────────────────────────────────
# Kata-kata ini bisa membentuk compound province/entity, sehingga match untuk
# label pendek TIDAK boleh diterima jika didahului kata-kata ini.
# Contoh: label="Riau" tidak boleh match "Kepulauan Riau".
_COMPOUND_GUARD = {
    "kepulauan", "dki", "di", "daerah", "nusa", "sulawesi",
    "kalimantan", "sumatera", "jawa", "barat", "timur", "tengah",
    "tenggara", "selatan", "utara",
}

_TOC_RE = re.compile(
    r"(\.\.\s*\d+|daftar isi|daftar tabel|daftar gambar|"
    r"table of contents|list of tables|\.\.\.\.+)",
    re.IGNORECASE,
)


# ─── Text helpers ──────────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).lower()).strip()


def label_match_strict(label: str, chunk_norm: str) -> bool:
    r"""
    True jika label ditemukan sebagai frasa utuh di chunk_norm.
    - Menggunakan word-boundary (?<!\w) / (?!\w) bukan string `in`.
    - Untuk label kata-tunggal, tolak match jika didahului compound prefix
      (contoh: 'Riau' tidak match 'Kepulauan Riau').
    - Untuk label multi-kata, langsung terima jika phrase match.
    """
    lnorm = normalize(label)
    if not lnorm or len(lnorm) < 2:
        return False

    pattern = r"(?<!\w)" + re.escape(lnorm) + r"(?!\w)"
    matches = list(re.finditer(pattern, chunk_norm))
    if not matches:
        return False

    # Multi-word label: phrase match is sufficient
    if " " in lnorm:
        return True

    # Single-word label: check each match for compound context
    for m in matches:
        before = chunk_norm[: m.start()].rstrip()
        prev_word = before.split()[-1] if before.split() else ""
        if prev_word.lower() not in _COMPOUND_GUARD:
            return True

    return False


def gold_value_variants(gv: str) -> list[str]:
    """Normalisation variants untuk angka gold_value."""
    gv = gv.strip().replace(" ", "")
    variants = {normalize(gv)}

    # Hapus titik sebagai ribuan separator: 22.573.548 → 22573548
    # Hanya jika ada pola .NNN (3 digit tepat setelah titik)
    no_dot = re.sub(r"\.(\d{3})(?=[.,\s]|$)", r"\1", gv)
    no_dot = re.sub(r"\.(\d{3})", r"\1", no_dot)
    if no_dot != gv:
        variants.add(normalize(no_dot))

    # Tukar separator desimal: 18,55 ↔ 18.55
    # Hanya untuk angka yang jelas desimal (≤2 digit setelah koma/titik)
    if "," in gv and "." not in gv:
        m = re.match(r'^(\d+),(\d{1,2})$', gv)
        if m:
            variants.add(normalize(gv.replace(",", ".")))
    elif "." in gv and "," not in gv:
        m = re.match(r'^(\d+)\.(\d{1,2})$', gv)
        if m:
            variants.add(normalize(gv.replace(".", ",")))

    # Hapus SEMUA separator hanya jika jelas ribuan (ada ≥2 titik/koma ribuan)
    _has_thousands = (
        len(re.findall(r"\.(\d{3})", gv)) >= 2 or
        len(re.findall(r",(\d{3})", gv)) >= 2
    )
    if _has_thousands:
        variants.add(normalize(re.sub(r"[.,]", "", gv)))

    return [v for v in variants if v and len(v) >= 2]


def is_toc_chunk(text: str) -> bool:
    hits = _TOC_RE.findall(text)
    return len(hits) >= 4 or (len(text) < 1500 and text.count("...") >= 3)


def word_overlap(reference: str, chunk_norm: str, min_len: int = 4,
                 threshold: float = 0.50) -> bool:
    words = [w for w in normalize(reference).split() if len(w) >= min_len]
    if not words:
        return False
    return (sum(1 for w in words if w in chunk_norm) / len(words)) >= threshold


def word_overlap_detail(reference: str, chunk_norm: str, min_len: int = 4,
                        threshold: float = 0.50) -> tuple:
    """Return (bool, matched_words_list, total_words_count) untuk reason yang lebih informatif."""
    words = [w for w in normalize(reference).split() if len(w) >= min_len]
    if not words:
        return False, [], 0
    matched = [w for w in words if w in chunk_norm]
    hit = (len(matched) / len(words)) >= threshold
    return hit, matched, len(words)


def extract_quote(text: str, signal: str, window: int = 280) -> str:
    """Kutipan konteks seputar sinyal yang ditemukan."""
    if not signal:
        snippet = text[:window].replace("\n", " ")
        return snippet if snippet else ""
    idx = text.lower().find(signal.lower())
    if idx == -1:
        return text[:window].replace("\n", " ")
    start = max(0, idx - window // 2)
    end   = min(len(text), idx + len(signal) + window // 2)
    snippet = text[start:end].replace("\n", " ")
    if start > 0:
        snippet = "\u2026" + snippet
    if end < len(text):
        snippet = snippet + "\u2026"
    return snippet


def verify_and_fix_candidate(row: dict, qa: dict) -> dict | None:
    """
    Re-verifikasi semua has_* flag dari chunk_text yang tersimpan.
    Update match_type, suggested_label, confidence, reason, evidence_quote.
    Return None jika tidak ada sinyal sama sekali (exclude dari output).
    """
    ctext = row.get("chunk_text", "")
    tn    = normalize(ctext)
    et    = str(qa.get("evidence_type", "") or "").strip()
    pg_s  = str(row.get("chunk_page_start", ""))
    pg_e  = str(row.get("chunk_page_end",   ""))
    ev    = str(qa.get("evidence_text",     "") or "").strip()
    anc   = str(qa.get("evidence_anchor",   "") or "").strip()
    pg    = str(qa.get("evidence_page_pdf", "") or "").strip()
    page_hit = page_overlaps(pg, pg_s, pg_e)

    if et == "table_row_column":
        gv  = str(qa.get("gold_value",   "") or "").strip()
        rl  = str(qa.get("row_label",    "") or "").strip()
        cl  = str(qa.get("column_label", "") or "").strip()
        tid = str(qa.get("table_id",     "") or "").strip()

        gv_v  = bool(gv  and any(label_match_strict(v, tn) for v in gold_value_variants(gv)))
        rl_v  = bool(rl  and label_match_strict(rl, tn))
        cl_v  = bool(cl  and label_match_strict(cl, tn))
        tid_v = bool(tid and label_match_strict(tid, tn))
        anc_v = bool(anc and label_match_strict(anc, tn))
        ev_v  = bool(ev  and normalize(ev) in tn)

        if gv_v and rl_v and cl_v:
            mt, sl, conf = "exact_table_evidence",   "2", "high"
        elif gv_v or (rl_v and cl_v and (not tid or tid_v)):
            mt, sl, conf = "partial_table_evidence",  "1", "medium"
        elif rl_v or cl_v or tid_v or anc_v or ev_v:
            mt, sl, conf = "keyword_only",            "0", "low"
        else:
            return None  # tidak ada sinyal → exclude

        score = (5 * gv_v + 3 * page_hit + 3 * rl_v +
                 3 * cl_v + 2 * tid_v + anc_v + ev_v)
        _anchor = gv if gv_v else (rl if rl_v else (cl if cl_v else ""))
        parts = []
        if gv_v:  parts.append(f'gold_value="{gv}"')
        if rl_v:  parts.append(f'row_label="{rl}"')
        if cl_v:  parts.append(f'col_label="{cl}"')
        if tid_v: parts.append(f'table_id="{tid}"')
        if anc_v: parts.append(f'anchor="{anc}"')
        if ev_v:  parts.append("evidence_text=found")
        if page_hit: parts.append(f"page_match={pg}")

        row.update({
            "has_gold_value":    gv_v,  "has_row_label":     rl_v,
            "has_column_label":  cl_v,  "has_table_id":      tid_v,
            "has_evidence_anchor": anc_v, "has_evidence_text": ev_v,
            "page_match":        page_hit,
            "match_type":        mt,  "suggested_label":   sl,
            "confidence":        conf, "strength_score":    score,
            "reason":            "; ".join(parts) if parts else "no_signal",
            "evidence_quote":    extract_quote(ctext, _anchor),
        })
        return row

    # ── Narrative / bullet / paragraph_table ────────────────────────────────
    ga   = str(qa.get("gold_answer", "") or "").strip()
    q    = str(qa.get("question",   "") or "").strip()
    ev_v = bool(ev and normalize(ev) in tn)
    ga_kw, ga_matched, ga_total = word_overlap_detail(ga, tn, min_len=4, threshold=0.50)
    anc_v  = bool(anc and label_match_strict(anc, tn))
    q_kw, q_matched, q_total   = word_overlap_detail(q,  tn, min_len=4, threshold=0.50)

    gv_v_pt = False
    if et == "paragraph_table":
        gv_pt = str(qa.get("gold_value", "") or "").strip()
        gv_v_pt = bool(gv_pt and any(label_match_strict(v, tn)
                                     for v in gold_value_variants(gv_pt)))

    score_n = (5 * ev_v + 3 * ga_kw + 2 * anc_v + 2 * q_kw +
               page_hit + gv_v_pt)
    if ev_v:
        mt, sl, conf = "exact_narrative_evidence",  "2", "high"
    elif ga_kw and anc_v:
        mt, sl, conf = "partial_narrative_context",  "1", "medium"
    elif ga_kw or anc_v or q_kw:
        mt, sl, conf = "keyword_only",               "0", "low"
    else:
        return None

    _anchor = ev[:60] if ev_v else (anc if anc_v else "")
    parts = []
    if ev_v:
        parts.append(f'evidence_text=found (len={len(ev)})')
    if ga_kw:
        ga_preview = ", ".join(ga_matched[:5])
        if len(ga_matched) > 5:
            ga_preview += f", ...+{len(ga_matched)-5}"
        parts.append(f'gold_answer_kw={len(ga_matched)}/{ga_total} [{ga_preview}]')
    if anc_v:
        parts.append(f'anchor="{anc}"')
    if q_kw:
        q_preview = ", ".join(q_matched[:4])
        if len(q_matched) > 4:
            q_preview += f", ...+{len(q_matched)-4}"
        parts.append(f'question_kw={len(q_matched)}/{q_total} [{q_preview}]')
    if page_hit:
        parts.append(f"page_match={pg}")
    if gv_v_pt:
        parts.append("gold_value_pt=found")

    row.update({
        "has_gold_value":     gv_v_pt, "has_row_label":      False,
        "has_column_label":   False,   "has_table_id":       False,
        "has_evidence_anchor": anc_v,  "has_evidence_text":  ev_v,
        "page_match":         page_hit,
        "match_type":         mt,  "suggested_label":    sl,
        "confidence":         conf, "strength_score":     score_n,
        "reason":             "; ".join(parts) if parts else "no_signal",
        "evidence_quote":     extract_quote(ctext, _anchor),
    })
    return row


def get_page_info(chunk: dict, method: str) -> tuple[str, str]:
    """Return (page_start, page_end) strings from chunk metadata."""
    meta = chunk.get("metadata", {})
    if method == "element_based":
        pr = str(meta.get("page_range", "")).strip()
        if not pr:
            pnums = meta.get("page_numbers", [])
            if pnums:
                pr = f"{min(pnums)}-{max(pnums)}"
        if "-" in pr:
            parts = pr.split("-", 1)
            return parts[0].strip(), parts[1].strip()
        return pr, pr
    else:
        pnums = meta.get("page_numbers", [])
        if pnums:
            return str(min(pnums)), str(max(pnums))
        return "", ""


def page_overlaps(ev_page: str, pg_start: str, pg_end: str) -> bool:
    """True jika evidence_page_pdf berada dalam range [pg_start, pg_end]."""
    if not ev_page or not pg_start:
        return False
    try:
        ep  = int(ev_page.strip())
        ps  = int(pg_start.strip())
        pe  = int(pg_end.strip()) if pg_end.strip() else ps
        return ps <= ep <= pe
    except ValueError:
        return False


# ─── Scoring ───────────────────────────────────────────────────────────────────

def score_chunk(chunk_text: str, pg_start: str, pg_end: str,
                qa: dict, method: str) -> dict:
    """
    Hitung evidence score dan flags untuk satu chunk vs satu query.

    Returns dict berisi:
      score, match_type, suggested_label, confidence, reason,
      evidence_quote, has_* flags
    """
    tn  = normalize(chunk_text)
    et  = str(qa.get("evidence_type", "") or "").strip()
    ev  = str(qa.get("evidence_text",  "") or "").strip()
    pg  = str(qa.get("evidence_page_pdf", "") or "").strip()
    anc = str(qa.get("evidence_anchor", "") or "").strip()

    page_hit = page_overlaps(pg, pg_start, pg_end)

    # ── TABLE QA ───────────────────────────────────────────────────────────────
    if et in ("table_row_column",):
        gv  = str(qa.get("gold_value",    "") or "").strip()
        rl  = str(qa.get("row_label",     "") or "").strip()
        cl  = str(qa.get("column_label",  "") or "").strip()
        tid = str(qa.get("table_id",      "") or "").strip()

        gv_found  = bool(gv  and any(label_match_strict(v, tn)
                                     for v in gold_value_variants(gv)))
        rl_found  = bool(rl  and label_match_strict(rl, tn))
        cl_found  = bool(cl  and label_match_strict(cl, tn))
        tid_found = bool(tid and label_match_strict(tid, tn))
        anc_found = bool(anc and label_match_strict(anc, tn))
        ev_found  = bool(ev  and normalize(ev) in tn)

        score = (
            (5 if gv_found  else 0) +
            (3 if page_hit  else 0) +
            (3 if rl_found  else 0) +
            (3 if cl_found  else 0) +
            (2 if tid_found else 0) +
            (1 if anc_found else 0) +
            (1 if ev_found  else 0)
        )

        # match_type determination (strict)
        if gv_found and rl_found and cl_found:
            match_type = "exact_table_evidence"
            suggested  = "2"
            confidence = "high"
        elif gv_found or (rl_found and cl_found and (not tid or tid_found)):
            match_type = "partial_table_evidence"
            suggested  = "1"
            confidence = "medium"
        elif rl_found or cl_found or tid_found or anc_found or ev_found:
            match_type = "keyword_only"
            suggested  = "0"
            confidence = "low"
        else:
            match_type = "not_relevant"
            suggested  = "0"
            confidence = "low"

        # Quote anchored on gold_value if found, else row/col
        _qa_anchor = (gv if gv_found else
                      (rl if rl_found else (cl if cl_found else "")))
        evidence_quote = extract_quote(chunk_text, _qa_anchor)

        parts = []
        if gv_found:  parts.append(f'gold_value="{gv}"')
        if rl_found:  parts.append(f'row_label="{rl}"')
        if cl_found:  parts.append(f'col_label="{cl}"')
        if tid_found: parts.append(f'table_id="{tid}"')
        if anc_found: parts.append(f'anchor="{anc}"')
        if ev_found:  parts.append("evidence_text=found")
        if page_hit:  parts.append(f"page_match={pg}")
        reason = "; ".join(parts) if parts else "no_signal"

        return {
            "score": score, "match_type": match_type,
            "suggested_label": suggested, "confidence": confidence,
            "reason": reason, "evidence_quote": evidence_quote,
            "has_gold_value": gv_found, "has_row_label": rl_found,
            "has_column_label": cl_found, "has_table_id": tid_found,
            "has_evidence_anchor": anc_found, "has_evidence_text": ev_found,
            "page_match": page_hit,
        }

    # ── NARRATIVE / BULLET / PARAGRAPH_TABLE ──────────────────────────────────
    ga = str(qa.get("gold_answer",    "") or "").strip()
    q  = str(qa.get("question",       "") or "").strip()

    ev_found              = bool(ev  and normalize(ev) in tn)
    ga_kw, ga_matched, ga_total = word_overlap_detail(ga, tn, min_len=4, threshold=0.50)
    anc_found             = bool(anc and label_match_strict(anc, tn))
    q_kw, q_matched, q_total   = word_overlap_detail(q,  tn, min_len=4, threshold=0.50)

    # For paragraph_table: also check table signals as secondary evidence
    gv_found_pt = False
    if et == "paragraph_table":
        gv  = str(qa.get("gold_value",   "") or "").strip()
        rl  = str(qa.get("row_label",    "") or "").strip()
        cl  = str(qa.get("column_label", "") or "").strip()
        gv_found_pt = bool(gv and any(label_match_strict(v, tn)
                                      for v in gold_value_variants(gv)))

    score = (
        (5 if ev_found    else 0) +
        (3 if ga_kw       else 0) +
        (2 if anc_found   else 0) +
        (2 if q_kw        else 0) +
        (1 if page_hit    else 0) +
        (1 if gv_found_pt else 0)
    )

    if ev_found:
        match_type = "exact_narrative_evidence"
        suggested  = "2"
        confidence = "high"
    elif ga_kw and anc_found:
        match_type = "partial_narrative_context"
        suggested  = "1"
        confidence = "medium"
    elif ga_kw or anc_found or q_kw:
        match_type = "keyword_only"
        suggested  = "0"
        confidence = "low"
    else:
        match_type = "not_relevant"
        suggested  = "0"
        confidence = "low"

    _quote_anchor = ev[:60] if ev_found else (anc if anc_found else "")
    evidence_quote = extract_quote(chunk_text, _quote_anchor)

    parts = []
    if ev_found:
        parts.append(f'evidence_text=found (len={len(ev)})')
    if ga_kw:
        matched_preview = ", ".join(ga_matched[:5])
        if len(ga_matched) > 5:
            matched_preview += f", ...+{len(ga_matched)-5}"
        parts.append(f'gold_answer_kw={len(ga_matched)}/{ga_total} [{matched_preview}]')
    if anc_found:
        parts.append(f'anchor="{anc}"')
    if q_kw:
        q_preview = ", ".join(q_matched[:4])
        if len(q_matched) > 4:
            q_preview += f", ...+{len(q_matched)-4}"
        parts.append(f'question_kw={len(q_matched)}/{q_total} [{q_preview}]')
    if page_hit:
        parts.append(f"page_match={pg}")
    if gv_found_pt:
        parts.append("gold_value_pt=found")
    reason = "; ".join(parts) if parts else "no_signal"

    return {
        "score": score, "match_type": match_type,
        "suggested_label": suggested, "confidence": confidence,
        "reason": reason, "evidence_quote": evidence_quote,
        "has_gold_value": gv_found_pt,
        "has_row_label": False, "has_column_label": False,
        "has_table_id": False,
        "has_evidence_anchor": anc_found, "has_evidence_text": ev_found,
        "page_match": page_hit,
    }


# ─── Load helpers ──────────────────────────────────────────────────────────────

def load_qa_gold() -> list[dict]:
    df = pd.read_excel(str(QA_XLSX), sheet_name="qa_gold", dtype=str).fillna("")
    rows = df.to_dict("records")
    print(f"[INFO] Loaded {len(rows)} QA gold dari {QA_XLSX.name}")
    return rows


def load_chunks(doc_id: str, method: str) -> list[dict] | None:
    stem = DOC_MAP.get(doc_id)
    if not stem:
        print(f"[WARN] doc_id tidak dikenal: {doc_id}", file=sys.stderr)
        return None
    fp = CHUNK_DIR / method / f"{stem}_chunks.json"
    if not fp.exists():
        print(f"[WARN] Chunk file tidak ditemukan: {fp}", file=sys.stderr)
        return None
    with open(fp, encoding="utf-8") as f:
        return json.load(f)


# ─── Candidate builder ─────────────────────────────────────────────────────────

_PRIORITY = {
    "exact_table_evidence": 4,
    "exact_narrative_evidence": 4,
    "partial_table_evidence": 3,
    "partial_narrative_context": 3,
    "keyword_only": 1,
    "not_relevant": 0,
}


def build_candidates_for_group(qa: dict, method: str,
                                pre_k: int, top_k: int) -> list[dict]:
    """
    Scan semua chunk, ambil pre_k terkuat, simpan top_k setelah re-rank.
    Semua dari JSON — tidak bergantung kandidat lama.
    """
    chunks = load_chunks(qa["doc_id"], method)
    if chunks is None:
        return []

    scored = []
    for chunk in chunks:
        cid    = str(chunk.get("chunk_id", ""))
        text   = chunk.get("text", "")
        pg_s, pg_e = get_page_info(chunk, method)

        if is_toc_chunk(text):
            continue

        res = score_chunk(text, pg_s, pg_e, qa, method)
        if res["score"] <= 0 or res["match_type"] == "not_relevant":
            continue

        scored.append({
            "chunk_id":       cid,
            "chunk_text":     text,
            "chunk_page_start": pg_s,
            "chunk_page_end":   pg_e,
            **res,
        })

    if not scored:
        return []

    # Sort: primary = score DESC, secondary = match_type priority DESC
    scored.sort(
        key=lambda x: (x["score"], _PRIORITY.get(x["match_type"], 0)),
        reverse=True,
    )

    # Pre-filter: top pre_k candidates
    pool = scored[:pre_k]

    # Final selection: prefer exact evidence; keep top_k
    exact = [c for c in pool if c["match_type"] in
             ("exact_table_evidence", "exact_narrative_evidence")]
    rest  = [c for c in pool if c not in exact]
    final = (exact + rest)[:top_k]

    # Re-verify all flags and fields from the actual chunk_text
    verified = []
    for c in final:
        fixed = verify_and_fix_candidate(c, qa)
        if fixed is not None:
            verified.append(fixed)
    return verified


# ─── Output builder ────────────────────────────────────────────────────────────

FIELDNAMES = [
    "query_id", "doc_id", "source_file", "question_preview", "evidence_type",
    "method", "chunk_id",
    "chunk_page_start", "chunk_page_end",
    "strength_score", "match_type", "suggested_label", "confidence",
    "has_gold_value", "has_row_label", "has_column_label",
    "has_table_id", "has_evidence_anchor", "has_evidence_text", "page_match",
    "evidence_quote", "reason",
    "label", "annotator", "status",
    "chunk_text_excerpt",
    "chunk_text",
]


def build_all_candidates(pre_k: int = PRE_K_DEFAULT,
                         top_k: int = TOP_K_DEFAULT) -> pd.DataFrame:
    qa_list = load_qa_gold()
    rows = []

    for qa in qa_list:
        qid    = qa["query_id"]
        doc_id = qa["doc_id"]
        et     = qa.get("evidence_type", "")
        sf     = qa.get("source_file",   "")
        q_prev = qa.get("question",      "")[:80]

        for method in METHODS:
            candidates = build_candidates_for_group(qa, method, pre_k, top_k)

            if not candidates:
                rows.append({
                    "query_id": qid, "doc_id": doc_id, "source_file": sf,
                    "question_preview": q_prev, "evidence_type": et,
                    "method": method, "chunk_id": "",
                    "chunk_page_start": "", "chunk_page_end": "",
                    "strength_score": 0,
                    "match_type": "not_found", "suggested_label": "0",
                    "confidence": "low",
                    "has_gold_value": False, "has_row_label": False,
                    "has_column_label": False, "has_table_id": False,
                    "has_evidence_anchor": False, "has_evidence_text": False,
                    "page_match": False,
                    "evidence_quote": "",
                    "reason": "no evidence-aware candidate found after full scan",
                    "label": "", "annotator": "",
                    "status": "needs_manual_validation",
                    "chunk_text_excerpt": "", "chunk_text": "",
                })
                continue

            for c in candidates:
                ctext   = c["chunk_text"]
                excerpt = ctext[:800].replace("\n", " ") if ctext else ""
                rows.append({
                    "query_id": qid, "doc_id": doc_id, "source_file": sf,
                    "question_preview": q_prev, "evidence_type": et,
                    "method": method, "chunk_id": c["chunk_id"],
                    "chunk_page_start": c["chunk_page_start"],
                    "chunk_page_end":   c["chunk_page_end"],
                    "strength_score":   c["score"],
                    "match_type":       c["match_type"],
                    "suggested_label":  c["suggested_label"],
                    "confidence":       c["confidence"],
                    "has_gold_value":   c["has_gold_value"],
                    "has_row_label":    c["has_row_label"],
                    "has_column_label": c["has_column_label"],
                    "has_table_id":     c["has_table_id"],
                    "has_evidence_anchor": c["has_evidence_anchor"],
                    "has_evidence_text":   c["has_evidence_text"],
                    "page_match":       c["page_match"],
                    "evidence_quote":   c["evidence_quote"],
                    "reason":           c["reason"],
                    "label": "", "annotator": "",
                    "status": "needs_manual_validation",
                    "chunk_text_excerpt": excerpt,
                    "chunk_text": ctext,
                })

    df = pd.DataFrame(rows, columns=FIELDNAMES)
    # Cast boolean flag columns
    for col in ["has_gold_value", "has_row_label", "has_column_label",
                "has_table_id", "has_evidence_anchor", "has_evidence_text",
                "page_match"]:
        df[col] = df[col].astype(bool)
    return df


# ─── Validation report ────────────────────────────────────────────────────────

def generate_validation_report(df: pd.DataFrame) -> str:
    """Post-generation integrity report: flags, quotes, match_type consistency."""
    lines = []
    sep = "-" * 62

    def w(s=""): lines.append(s)

    w("=" * 62)
    w("VALIDATION REPORT — KANDIDAT V3")
    w("=" * 62)
    w()
    w(f"  Total baris           : {len(df)}")
    w(f"  Total grup            : {df.groupby(['query_id','method']).ngroups}")
    w()
    w("  match_type distribution:")
    for mt, cnt in df["match_type"].value_counts().items():
        w(f"    {mt:<35}: {cnt}")
    w()
    w("  suggested_label distribution:")
    for sl, cnt in df["suggested_label"].value_counts().sort_index().items():
        w(f"    label {sl}: {cnt}")
    w()

    exact = df[df["match_type"].isin(
        ["exact_table_evidence", "exact_narrative_evidence"])]
    w(f"  Exact evidence total  : {len(exact)}")

    # exact_table: verify gv+rl+cl all True
    et_rows = df[df["match_type"] == "exact_table_evidence"]
    valid_et = et_rows[
        (et_rows["has_gold_value"]   == True) &
        (et_rows["has_row_label"]    == True) &
        (et_rows["has_column_label"] == True)
    ]
    invalid_et = len(et_rows) - len(valid_et)
    w(f"  exact_table valid (gv+rl+cl) : {len(valid_et)} / {len(et_rows)}")
    if invalid_et:
        w(f"  [WARN] exact_table dengan flag tidak lengkap: {invalid_et}")
        bad = et_rows[
            ~((et_rows["has_gold_value"] == True) &
              (et_rows["has_row_label"]  == True) &
              (et_rows["has_column_label"] == True))
        ]
        for _, r in bad.iterrows():
            w(f"    {r['query_id']} / {r['method']} / chunk {r['chunk_id']}")
    w()

    # evidence_quote mismatch: quote not in chunk_text
    mismatch_quote = 0
    mismatch_rows  = []
    for _, r in df.iterrows():
        eq   = str(r.get("evidence_quote", "") or "").strip()
        ct   = str(r.get("chunk_text",     "") or "").strip()
        if not eq or not ct:
            continue
        clean_eq = eq.lstrip("…").rstrip("…").strip()
        if clean_eq and normalize(clean_eq[:60]) not in normalize(ct):
            mismatch_quote += 1
            mismatch_rows.append(
                f"    {r['query_id']} / {r['method']} / chunk {r['chunk_id']}")
    w(f"  evidence_quote mismatch: {mismatch_quote}")
    for mr in mismatch_rows[:10]:
        w(mr)
    if len(mismatch_rows) > 10:
        w(f"    ... dan {len(mismatch_rows) - 10} lainnya")
    w()

    # Groups without any exact evidence
    all_groups = set(df.groupby(["query_id","method"]).groups.keys())
    exact_groups = set(
        df[df["match_type"].isin(
            ["exact_table_evidence","exact_narrative_evidence"]
        )].apply(lambda r: (r["query_id"], r["method"]), axis=1)
    )
    no_exact = sorted(all_groups - exact_groups)
    w(f"  Grup tanpa exact evidence: {len(no_exact)}")
    for g in no_exact:
        sub = df[(df["query_id"]==g[0]) & (df["method"]==g[1])]
        best_mt = sub["match_type"].iloc[0] if len(sub) else "empty"
        w(f"    {g[0]} × {g[1]:<22} best={best_mt}")
    w()

    # Narrative QA queries without exact evidence (query level, not group level)
    w(sep)
    w("NARASI — QUERY TANPA EXACT EVIDENCE")
    w(sep)
    narrative_types = ["paragraph", "bullet", "paragraph_table"]
    narr_df = df[df["evidence_type"].isin(narrative_types)]
    narr_exact_qids = set(
        narr_df[narr_df["match_type"] == "exact_narrative_evidence"]["query_id"]
    )
    all_narr_qids = set(narr_df["query_id"].unique())
    narr_no_exact = sorted(all_narr_qids - narr_exact_qids)
    w(f"  Query narasi total            : {len(all_narr_qids)}")
    w(f"  Query narasi dg exact evidence: {len(narr_exact_qids)}")
    w(f"  Query narasi tanpa exact       : {len(narr_no_exact)}")
    for qid in narr_no_exact:
        best_row = narr_df[narr_df["query_id"] == qid].sort_values(
            "strength_score", ascending=False).iloc[0]
        w(f"    {qid}  best_match={best_row['match_type']}  method={best_row['method']}")
    w()

    # Candidates entered only because of question_kw
    w(sep)
    w("KANDIDAT YANG MASUK HANYA KARENA question_kw")
    w(sep)
    qkw_only = df[
        df["reason"].str.contains("question_kw=", na=False) &
        ~df["reason"].str.contains("evidence_text=found", na=False) &
        ~df["reason"].str.contains("gold_answer_kw=", na=False) &
        ~df["reason"].str.contains("gold_value=\"", na=False) &
        ~df["reason"].str.contains("row_label=\"", na=False) &
        ~df["reason"].str.contains("col_label=\"", na=False)
    ]
    w(f"  Total kandidat masuk hanya question_kw: {len(qkw_only)}")
    for _, r in qkw_only.iterrows():
        w(f"    {r['query_id']} / {r['method']} / chunk {r['chunk_id']}  mt={r['match_type']}")
    w()

    # label must be blank
    w(sep)
    w("INTEGRITAS LABEL")
    w(sep)
    non_blank_labels = df[df["label"].astype(str).str.strip() != ""]
    if len(non_blank_labels) == 0:
        w("  label semua KOSONG ✓")
    else:
        w(f"  [WARN] {len(non_blank_labels)} baris memiliki label tidak kosong!")
        for _, r in non_blank_labels.iterrows():
            w(f"    {r['query_id']} / {r['method']} / chunk {r['chunk_id']}  label={r['label']}")
    w()

    # status must be needs_manual_validation
    non_nmv = df[df["status"].astype(str).str.strip() != "needs_manual_validation"]
    if len(non_nmv) == 0:
        w("  status semua 'needs_manual_validation' ✓")
    else:
        w(f"  [WARN] {len(non_nmv)} baris status tidak sesuai:")
        for _, r in non_nmv.iterrows():
            w(f"    {r['query_id']} / {r['method']} / chunk {r['chunk_id']}  status={r['status']}")
    w()

    w(sep)
    return "\n".join(lines)


# ─── Summary builder ───────────────────────────────────────────────────────────

def build_summary(df: pd.DataFrame, out_dir: Path,
                  pre_k: int, top_k: int) -> str:
    lines = []
    sep = "=" * 62

    def w(s=""):
        lines.append(s)

    w(sep)
    w("RINGKASAN KANDIDAT V3 — EVIDENCE-AWARE")
    w(sep)
    w()

    total      = len(df)
    not_found  = (df["match_type"] == "not_found").sum()
    active     = total - not_found
    n_exact_t  = (df["match_type"] == "exact_table_evidence").sum()
    n_exact_n  = (df["match_type"] == "exact_narrative_evidence").sum()
    n_partial  = df["match_type"].isin(
        ["partial_table_evidence", "partial_narrative_context"]).sum()
    n_kw       = df["match_type"].isin(["keyword_only", "not_relevant"]).sum()
    n_sug2     = (df["suggested_label"] == "2").sum()
    n_sug1     = (df["suggested_label"] == "1").sum()

    w(f"  Pre-filter k          : {pre_k}")
    w(f"  Final top-k           : {top_k}")
    w()
    w(f"  Total baris           : {total}")
    w(f"  not_found             : {not_found}")
    w(f"  Aktif (ada chunk)     : {active}")
    w()
    w(f"  Saran label 2 (exact) : {n_sug2}")
    w(f"  Saran label 1 (partial): {n_sug1}")
    w()

    n_groups = df.groupby(["query_id", "method"]).ngroups
    w(f"  Grup query×method     : {n_groups}  (target: 90)")
    avg = active / max(n_groups - not_found, 1)
    w(f"  Rata-rata per grup    : {active / max(n_groups, 1):.2f}")
    w()

    w("  match_type breakdown:")
    for mt, cnt in df["match_type"].value_counts().items():
        w(f"    {mt:<35}: {cnt}")
    w()

    w("  Per method:")
    for m in METHODS:
        sub = df[(df["method"] == m) & (df["match_type"] != "not_found")]
        nf  = (df[(df["method"] == m)]["match_type"] == "not_found").sum()
        w(f"    {m:<22}: {len(sub):>3} kandidat, {nf} not_found")
    w()

    # Grup dengan minimal 1 exact evidence
    exact_grps = df[df["match_type"].isin(
        ["exact_table_evidence", "exact_narrative_evidence"]
    )].groupby(["query_id", "method"]).ngroups
    w(f"  Grup dg ≥1 exact evidence : {exact_grps}")
    w()

    # ── Comparison with v2 ────────────────────────────────────────────────────
    w(sep)
    w("PERBANDINGAN DENGAN V2")
    w(sep)

    if V2_XLSX.exists():
        try:
            v2 = pd.read_excel(str(V2_XLSX), sheet_name="candidates", dtype=str).fillna("")
            v2_not_found = (v2["match_type"] == "not_found").sum()
            v2_active    = len(v2) - v2_not_found
            v2_exact     = v2["match_type"].isin(
                ["exact_table_evidence", "exact_narrative_evidence"]).sum()
            v2_grps      = v2.groupby(["query_id", "method"]).ngroups

            w()
            w(f"  {'Metrik':<30} {'v2':>8} {'v3':>8}  {'Delta':>8}")
            w(f"  {'-'*56}")
            metrics = [
                ("Total baris",        len(v2),        total),
                ("not_found",          v2_not_found,   not_found),
                ("Aktif kandidat",     v2_active,      active),
                ("Exact evidence",     v2_exact,       n_exact_t + n_exact_n),
                ("Saran label 2",
                 (v2["suggested_label"] == "2").sum(), n_sug2),
                ("Saran label 1",
                 (v2["suggested_label"] == "1").sum(), n_sug1),
                ("Total grup",         v2_grps,        n_groups),
            ]
            for name, v2_val, v3_val in metrics:
                delta = v3_val - v2_val
                sign  = "+" if delta > 0 else ""
                w(f"  {name:<30} {v2_val:>8} {v3_val:>8}  {sign}{delta:>7}")
            w()

            # Grup yang membaik (ada exact di v3, tidak di v2)
            v3_exact_grps = set(
                df[df["match_type"].isin(
                    ["exact_table_evidence", "exact_narrative_evidence"]
                )].apply(lambda r: (r["query_id"], r["method"]), axis=1)
            )
            v2_exact_grps = set(
                v2[v2["match_type"].isin(
                    ["exact_table_evidence", "exact_narrative_evidence"]
                )].apply(lambda r: (r["query_id"], r["method"]), axis=1)
            )
            improved = v3_exact_grps - v2_exact_grps
            regressed = v2_exact_grps - v3_exact_grps

            w(f"  Grup membaik (v3 punya exact, v2 tidak) : {len(improved)}")
            for g in sorted(improved)[:10]:
                w(f"    {g[0]} × {g[1]}")
            if len(improved) > 10:
                w(f"    ... dan {len(improved) - 10} grup lainnya")
            w()
            w(f"  Grup masih partial/keyword only          : "
              f"{n_groups - exact_grps}")
            w(f"  Grup not_found                           : {not_found}")

        except Exception as e:
            w(f"  [WARN] Gagal baca v2 untuk perbandingan: {e}")
    else:
        w("  [INFO] File v2 tidak ditemukan. Perbandingan dilewati.")

    w()
    w(sep)
    w("CATATAN PENTING")
    w(sep)
    w()
    w("  - suggested_label adalah REKOMENDASI AUDIT — bukan label final")
    w("  - label KOSONG — harus diisi manual oleh peneliti")
    w("  - Evaluasi retrieval: Precision@k, Recall@k, MRR")
    w("  - word-boundary matching aktif — mencegah false positive label pendek")
    w("  - 'Riau' tidak akan match 'Kepulauan Riau' secara otomatis")
    w()

    return "\n".join(lines)


# ─── Save ──────────────────────────────────────────────────────────────────────

def save_outputs(df: pd.DataFrame, summary_text: str, out_dir: Path,
                 validation_text: str = "") -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    # CSV
    csv_path = out_dir / f"{OUT_STEM}.csv"
    df.to_csv(str(csv_path), index=False, encoding="utf-8-sig")
    print(f"\n[OK] CSV  → {csv_path}")

    # XLSX
    xlsx_path = out_dir / f"{OUT_STEM}.xlsx"
    with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="candidates", index=False)
        ws = writer.sheets["candidates"]

        hdr_fill = PatternFill("solid", fgColor="1E293B")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color-code suggested_label
        sl_col = next(
            (i for i, c in enumerate(ws[1], 1) if c.value == "suggested_label"),
            None,
        )
        if sl_col:
            colors = {"2": "D1FAE5", "1": "FEF3C7", "0": "FEE2E2"}
            for row in ws.iter_rows(min_row=2):
                sl_cell = row[sl_col - 1]
                fill_color = colors.get(str(sl_cell.value or ""), "FFFFFF")
                sl_cell.fill = PatternFill("solid", fgColor=fill_color)

        # Color-code match_type
        mt_col = next(
            (i for i, c in enumerate(ws[1], 1) if c.value == "match_type"),
            None,
        )
        mt_colors = {
            "exact_table_evidence":    "DCFCE7",
            "exact_narrative_evidence":"DCFCE7",
            "partial_table_evidence":  "FEF9C3",
            "partial_narrative_context":"FEF9C3",
            "keyword_only":            "FEE2E2",
            "not_relevant":            "F1F5F9",
            "not_found":               "E2E8F0",
        }
        if mt_col:
            for row in ws.iter_rows(min_row=2):
                mt_cell = row[mt_col - 1]
                fc = mt_colors.get(str(mt_cell.value or ""), "FFFFFF")
                mt_cell.fill = PatternFill("solid", fgColor=fc)

        ws.freeze_panes = "A2"

        col_widths = {
            "query_id": 9, "doc_id": 12, "source_file": 28,
            "question_preview": 40, "evidence_type": 16, "method": 18,
            "chunk_id": 10, "chunk_page_start": 12, "chunk_page_end": 12,
            "strength_score": 13, "match_type": 26, "suggested_label": 14,
            "confidence": 11,
            "has_gold_value": 13, "has_row_label": 13,
            "has_column_label": 14, "has_table_id": 12,
            "has_evidence_anchor": 16, "has_evidence_text": 15,
            "page_match": 11,
            "evidence_quote": 52, "reason": 44,
            "label": 12, "annotator": 13, "status": 22,
            "chunk_text_excerpt": 65, "chunk_text": 80,
        }
        for col_num, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_num)
            ].width = col_widths.get(col_name, 14)

    print(f"[OK] XLSX → {xlsx_path}")

    # Summary TXT
    txt_path = out_dir / "summary_candidate_v3_evidence_aware.txt"
    txt_path.write_text(summary_text, encoding="utf-8")
    print(f"[OK] TXT  → {txt_path}")

    # Validation TXT (separate file)
    if validation_text:
        val_path = out_dir / "validation_candidate_v3_after_narrative_revision.txt"
        val_path.write_text(validation_text, encoding="utf-8")
        print(f"[OK] TXT  → {val_path}")


# ─── Validation ────────────────────────────────────────────────────────────────

def validate(df: pd.DataFrame) -> None:
    print("\n[VALIDASI]")
    n_groups = df.groupby(["query_id", "method"]).ngroups
    ok_groups = n_groups == 90
    print(f"  Jumlah grup : {n_groups}  {'[OK]' if ok_groups else '[WARN] target 90'}")

    orphan = []
    for (qid, m), g in df.groupby(["query_id", "method"]):
        if len(g) == 0:
            orphan.append(f"{qid}/{m}")
    if orphan:
        print(f"  [WARN] Grup kosong: {orphan}")
    else:
        print("  Semua grup punya >=1 baris (kandidat atau placeholder) [OK]")

    over = df.groupby(["query_id", "method"]).size()
    over_k = (over > TOP_K_DEFAULT).sum()
    if over_k:
        print(f"  [WARN] {over_k} grup memiliki lebih dari {TOP_K_DEFAULT} kandidat")

    print(f"  not_found             : {(df['match_type'] == 'not_found').sum()}")
    print(f"  exact evidence grup   : "
          f"{df[df['match_type'].isin(['exact_table_evidence','exact_narrative_evidence'])].groupby(['query_id','method']).ngroups}")


# ─── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(
        description="Build evidence-aware retrieval candidates v3",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--top_k",  type=int, default=TOP_K_DEFAULT,
                        help=f"Kandidat final per grup (default: {TOP_K_DEFAULT})")
    parser.add_argument("--pre_k",  type=int, default=PRE_K_DEFAULT,
                        help=f"Pre-filter per grup sebelum top_k (default: {PRE_K_DEFAULT})")
    parser.add_argument("--output_dir", type=str, default=str(OUT_DIR),
                        help=f"Direktori output (default: {OUT_DIR})")
    args = parser.parse_args()

    print(f"\n[INFO] Build candidate v3 — pre_k={args.pre_k}, top_k={args.top_k}")
    print(f"[INFO] QA gold : {QA_XLSX}")
    print(f"[INFO] Chunks  : {CHUNK_DIR}")
    print()

    df = build_all_candidates(pre_k=args.pre_k, top_k=args.top_k)
    validate(df)

    validation = generate_validation_report(df)
    print("\n" + validation)

    summary = build_summary(df, Path(args.output_dir), args.pre_k, args.top_k)
    print("\n" + summary)

    save_outputs(df, summary + "\n\n" + validation, Path(args.output_dir),
                 validation_text=validation)

    print("\nLANGKAH SELANJUTNYA:")
    print("  1. Restart Streamlit agar file v3 terdeteksi sebagai kandidat aktif:")
    print("     streamlit run src/streamlit/app.py")
    print("  2. Tinjau suggested_label dan has_* flags sebagai panduan")
    print("  3. Isi label secara manual: 1=relevan, 0=tidak relevan")


if __name__ == "__main__":
    main()
